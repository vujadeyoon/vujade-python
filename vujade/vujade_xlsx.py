"""
Dveloper: vujadeyoon
Email: vujadeyoon@gmail.com
Github: https://github.com/vujadeyoon/vujade

Title: vujade_xlsx.py
Description: A module for xlsx
"""

import openpyxl
import pandas as pd
import numpy as np
from pandas.io.formats.style import Styler
from typing import Optional
from vujade import vujade_path as path_


class XLSX(object):
    def __init__(self, _spath_filename: str, _mode: str) -> None:
        super(XLSX, self).__init__()
        self.spath_filename = _spath_filename
        self.mode = _mode

        if _mode == 'w':
            self.wb = openpyxl.Workbook()
        elif _mode == 'r' or _mode == 'a':
            self.wb = openpyxl.load_workbook(filename=self.spath_filename)
        else:
            raise NotImplementedError('The _mode, {} may be incorrect.'.format(_mode))

        self.ws = self.wb.active

    def create_sheet(self, _title: Optional[str] = None, _index: Optional[int] = None) -> openpyxl.Workbook.worksheets:
        return self.wb.create_sheet(title=_title, index=_index)

    def save_workbook(self) -> None:
        self.wb.save(self.spath_filename)


class Ndarr2XLSX(object):
    def __init__(self, _spath_filename: str, _is_remove: bool = False, _header: Optional[list] = None, _index: bool = False):
        super(Ndarr2XLSX, self).__init__()
        self.path_filename = path_.Path(_spath=_spath_filename)
        self.header = _header
        self.index = _index

        if _is_remove is True and self.path_filename.path.is_file() is True:
            self.path_filename.path.unlink()

    def write(self, _ndarr: np.array, _attribute: Optional[list] = None) -> None:
        """
        :param _attribute: list of the attributes,
                           _attribute[0] = {
                               'idx_cell': (1, 1),        # Required
                               'background-color': 'red', # Option
                               'color': 'violet'          # Option
                           }
        """
        if _attribute is not None:
            is_styler = True
        else:
            is_styler = False

        if is_styler is True:
            if self.path_filename.ext != '.xlsx':
                raise ValueError('The extension of the file should be .xlsx, not {} because of the styling.'.format(self.path_filename.ext))

        df = pd.DataFrame(_ndarr)
        if is_styler is True:
            styler = df.style.apply(self._styling_cell, _attribute=_attribute, axis = None)
            styler.to_excel(self.path_filename.str, engine='openpyxl', header=self.header, index=self.index)
        else:
            df.to_excel(self.path_filename.str, engine='openpyxl', header=self.header, index=self.index)

    def _styling_cell(self, _x, _attribute: list) -> Styler:
        styler = pd.DataFrame('', index=_x.index, columns=_x.columns)

        for _idx, (_att) in enumerate(_attribute):
            if not 'idx_cell' in _att.keys():
                raise ValueError('The key value of the attribute, idx_cell should be required.')

            styler_code = ''
            for _idx, (_att_key, _att_val) in enumerate(_att.items()):
                if _att_key == 'idx_cell':
                    idx_row, idx_col = _att_val
                else:
                    styler_code += '{}: {}'.format(_att_key, _att_val)
                    if _idx < (len(_attribute) - 1):
                        styler_code += '; '

            styler.iloc[idx_row, idx_col] = styler_code

        return styler


def csv2xlsx(_spath_src: str, _spath_dst: str, _header: bool = True, _index: bool = True) -> None:
    df = pd.read_csv(_spath_src)
    df.to_excel(_spath_dst, header=_header, index=_index)


def worksheet2df(_ws: openpyxl.Workbook.worksheets, _is_header: bool = True) -> pd.DataFrame:
    ws_values = _ws.values

    if _is_header is True:
        header = next(ws_values)[0:]
        res = pd.DataFrame(ws_values, columns=header)
    else:
        res = pd.DataFrame(ws_values)

    return res
